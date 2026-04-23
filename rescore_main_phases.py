#!/usr/bin/env python3
"""
Re-runs S4 queries for Phase 1 and Phase 2 main result files (90 questions each),
then rescores ALL strategies using doc-level hit detection from retrieved_doc_ids.

Adds source_doc and retrieved_doc_ids to the stored results.
Checkpoint-based — safe to interrupt and resume with --resume.
"""

import json, time, requests, sys, argparse, openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

API_BASE = "http://127.0.0.1:8007"
DATA_DIR = Path("/www/wwwroot/chunkbench.tech/data")
RESULTS  = Path("/www/wwwroot/chunkbench.tech/data/results")

QUESTION_FILE = DATA_DIR / "questions_v1_FROZEN.json"
PHASE1_FILE   = RESULTS / "phase1" / "phase1_final.json"
PHASE2_FILE   = RESULTS / "phase2" / "phase2_final.json"
CHECKPOINT    = RESULTS / "rescore_main_checkpoint.json"

STRATS       = ["S1", "S2", "S3", "S4", "B0"]
QUERY_PARAMS = {
    "k": 10, "context_mode": "fixed-budget",
    "context_budget_tokens": 1800, "temperature": 0.0,
    "max_tokens": 512, "dedup": True,
}

PHASE1_CFG = {"S1":"size=512,overlap=50(10%)","S2":"size=512,overlap=50(10%)",
              "S3":"percentile/95","S4":"proposition/default","B0":"no retrieval"}
PHASE2_CFG = {"S1":"size=512,overlap=102(20%)","S2":"size=1024,overlap=102(10%)",
              "S3":"threshold=85,max=2000","S4":"proposition/default","B0":"no retrieval"}


def _norm(d):
    if not d: return d
    if d.endswith(".pdf"): d = d[:-4]
    return d.replace("_shallow","").replace("_deep","")

def doc_hit(source_doc, doc_ids, k):
    if not source_doc or not doc_ids: return 0
    src = _norm(source_doc)
    return int(any(_norm(x) == src for x in doc_ids[:k]))

def doc_mrr(source_doc, doc_ids):
    if not source_doc or not doc_ids: return 0.0
    src = _norm(source_doc)
    for i, x in enumerate(doc_ids):
        if _norm(x) == src: return 1.0 / (i + 1)
    return 0.0

def call_query(question, mode, strategies):
    r = requests.post(f"{API_BASE}/query",
                      json={"question": question, "mode": mode,
                            "strategies": strategies, **QUERY_PARAMS},
                      timeout=300)
    r.raise_for_status()
    return r.json()["results"]


def rescore_from_doc_ids(entry, source_doc):
    """Re-score all strategies in entry using doc-level hit from retrieved_doc_ids."""
    for sid, sr in entry.get("strategies", {}).items():
        if sid == "B0":
            sr["hit_at_1"] = None
            sr["hit_at_3"] = None
            sr["hit_at_5"] = None
            sr["mrr"]      = None
            continue
        doc_ids = sr.get("retrieved_doc_ids", [])
        if not doc_ids: continue
        sr["hit_at_1"] = doc_hit(source_doc, doc_ids, 1)
        sr["hit_at_3"] = doc_hit(source_doc, doc_ids, 3)
        sr["hit_at_5"] = doc_hit(source_doc, doc_ids, 5)
        sr["mrr"]      = doc_mrr(source_doc, doc_ids)


def run(resume=False):
    questions  = {q["q_id"]: q for q in json.loads(QUESTION_FILE.read_text())}
    phase1_map = {r["q_id"]: r for r in json.loads(PHASE1_FILE.read_text())}
    phase2_map = {r["q_id"]: r for r in json.loads(PHASE2_FILE.read_text())}

    done_ids = set()
    if resume and CHECKPOINT.exists():
        cp = json.loads(CHECKPOINT.read_text())
        done_ids = set(cp.get("done_ids", []))
        # restore updated entries
        for qid, entry in cp.get("phase1", {}).items():
            phase1_map[qid] = entry
        for qid, entry in cp.get("phase2", {}).items():
            phase2_map[qid] = entry
        print(f"Resumed: {len(done_ids)} already done")

    all_ids = sorted(set(phase1_map) | set(phase2_map))
    remaining = [qid for qid in all_ids if qid not in done_ids]
    total = len(all_ids)

    print(f"Total: {total} | Remaining: {len(remaining)}")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    for n, qid in enumerate(remaining, len(done_ids)+1):
        q = questions.get(qid)
        if not q:
            print(f"[{n}/{total}] {qid}: not in question file, skipping")
            done_ids.add(qid)
            continue

        source_doc = q.get("source_doc", "")
        print(f"[{n:03d}/{total}] {qid}  source={source_doc}", flush=True)
        t0 = time.time()

        for attempt in range(3):
            try:
                # Phase 1 — query S4 only (S1/S2/S3 already have top_chunks with doc_ids)
                p1 = call_query(q["question"], "phase1", ["S1","S2","S3","S4"])
                p2 = call_query(q["question"], "phase2", ["S1","S2","S3","S4"])

                for phase_map, phase_res, mode_label in [
                    (phase1_map, p1, "phase1"),
                    (phase2_map, p2, "phase2"),
                ]:
                    entry = phase_map.get(qid, {"q_id": qid, "question": q["question"],
                                                 "category": q.get("category"),
                                                 "needle_type": q.get("needle_type"),
                                                 "gold_span": q.get("gold_span"),
                                                 "gold_answer": q.get("gold_answer"),
                                                 "source_doc": source_doc,
                                                 "strategies": {}})
                    entry["source_doc"] = source_doc

                    for sid, sr in phase_res.items():
                        chunks   = sr.get("retrieved_chunks", [])
                        doc_ids  = list(dict.fromkeys(c["doc_id"] for c in chunks))
                        top3     = [{"rank": c["rank"], "doc_id": c["doc_id"],
                                     "distance": round(c.get("distance",0),4),
                                     "text": c["text"][:400]} for c in chunks[:3]]
                        entry["strategies"][sid] = {
                            "answer":            sr.get("answer",""),
                            "latency_s":         sr.get("latency_s",0),
                            "context_tokens":    sr.get("context_tokens_est",0),
                            "retrieved_doc_ids": doc_ids,
                            "top_chunks":        top3,
                            "hit_at_1": doc_hit(source_doc, doc_ids, 1),
                            "hit_at_3": doc_hit(source_doc, doc_ids, 3),
                            "hit_at_5": doc_hit(source_doc, doc_ids, 5),
                            "mrr":      doc_mrr(source_doc, doc_ids),
                        }
                    phase_map[qid] = entry

                elapsed = round(time.time() - t0, 1)
                h = {s: phase2_map[qid]["strategies"].get(s,{}).get("hit_at_5","?")
                     for s in ["S1","S2","S3","S4"]}
                print(f"  ok ({elapsed}s) | " + " ".join(f"{s}:{v}" for s,v in h.items()), flush=True)
                break
            except Exception as e:
                if attempt < 2: time.sleep(15)
                else: print(f"  SKIP: {e}", flush=True)

        done_ids.add(qid)
        CHECKPOINT.write_text(json.dumps({
            "done_ids": list(done_ids),
            "phase1": {k: phase1_map[k] for k in done_ids if k in phase1_map},
            "phase2": {k: phase2_map[k] for k in done_ids if k in phase2_map},
        }, ensure_ascii=False))
        time.sleep(4)

    return list(phase1_map.values()), list(phase2_map.values())


def summary(out, label):
    print(f"\n--- {label} ---")
    for sid in STRATS:
        scored = [r for r in out
                  if r.get("strategies",{}).get(sid,{}).get("hit_at_5") is not None
                  and sid != "B0"]
        if not scored: continue
        h5  = [r["strategies"][sid]["hit_at_5"] for r in scored]
        h1  = [r["strategies"][sid]["hit_at_1"] for r in scored]
        mrr = [r["strategies"][sid]["mrr"]       for r in scored]
        print(f"  {sid}: Hit@1={round(sum(h1)/len(h1)*100,1)}%  "
              f"Hit@5={round(sum(h5)/len(h5)*100,1)}%  "
              f"MRR={round(sum(mrr)/len(mrr),3)}  n={len(h5)}")


GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BOLD   = Font(bold=True)


def make_excel(out, path, label, cfg):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Summary"
    ws.append([label])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append(["Hit detection: doc-level (source_doc in top-k retrieved doc_ids)"])
    ws.append([])
    ws.append(["Strategy","Config","N","Hit@1(%)","Hit@3(%)","Hit@5(%)","MRR"])
    for cell in ws[5]: cell.font = BOLD
    for sid in STRATS:
        scored = [r for r in out
                  if r.get("strategies",{}).get(sid,{}).get("hit_at_5") is not None
                  and sid != "B0"]
        if not scored:
            ws.append([sid, cfg.get(sid,""), 0, None, None, None, None]); continue
        h5  = [r["strategies"][sid]["hit_at_5"] for r in scored]
        h3  = [r["strategies"][sid]["hit_at_3"] for r in scored]
        h1  = [r["strategies"][sid]["hit_at_1"] for r in scored]
        m   = [r["strategies"][sid]["mrr"]       for r in scored]
        v5  = round(sum(h5)/len(h5)*100, 1)
        ws.append([sid, cfg.get(sid,""), len(h5),
                   round(sum(h1)/len(h1)*100,1), round(sum(h3)/len(h3)*100,1),
                   v5, round(sum(m)/len(m),3)])
        ws.cell(ws.max_row,6).fill = GREEN if v5>=40 else RED if v5<15 else YELLOW

    ws2 = wb.create_sheet("Per-Question")
    hdr = ["q_id","source_doc","category","question","gold_span","gold_answer"]
    for sid in STRATS: hdr += [f"{sid}_hit5",f"{sid}_hit3",f"{sid}_mrr",f"{sid}_docs"]
    for sid in STRATS: hdr.append(f"{sid}_answer")
    ws2.append(hdr)
    for cell in ws2[1]: cell.font = BOLD
    for r in out:
        row = [r["q_id"], r.get("source_doc",""), r.get("category",""),
               r.get("question","")[:150], r.get("gold_span","")[:150],
               r.get("gold_answer","")[:100]]
        for sid in STRATS:
            sr   = r.get("strategies",{}).get(sid,{})
            docs = ", ".join(sr.get("retrieved_doc_ids",[]))
            row += [sr.get("hit_at_5"), sr.get("hit_at_3"), sr.get("mrr"), docs]
        for sid in STRATS:
            row.append(r.get("strategies",{}).get(sid,{}).get("answer","")[:250])
        ws2.append(row)
        rn = ws2.max_row
        for i, h in enumerate(hdr):
            if h.endswith("_hit5"):
                c = ws2.cell(rn, i+1)
                if c.value == 1: c.fill = GREEN
                elif c.value == 0: c.fill = RED

    for s in [ws, ws2]:
        for col in s.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            s.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2,45)
    wb.save(str(path))
    print(f"  Excel: {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true")
    args = parser.parse_args()

    try:
        r = requests.get(f"{API_BASE}/health", timeout=10)
        print(f"API: {r.json()}")
    except Exception as e:
        print(f"API not reachable: {e}"); sys.exit(1)

    p1_out, p2_out = run(resume=args.resume)

    PHASE1_FILE.write_text(json.dumps(p1_out, indent=2, ensure_ascii=False))
    PHASE2_FILE.write_text(json.dumps(p2_out, indent=2, ensure_ascii=False))
    print("\nJSON saved.")

    summary(p1_out, "Phase 1")
    summary(p2_out, "Phase 2")

    make_excel(p1_out, RESULTS/"phase1"/"phase1_results.xlsx", "Phase 1 — 90 questions", PHASE1_CFG)
    make_excel(p2_out, RESULTS/"phase2"/"phase2_results.xlsx", "Phase 2 — 90 questions", PHASE2_CFG)

    print(f"\nDone: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
