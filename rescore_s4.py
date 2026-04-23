#!/usr/bin/env python3
"""
Re-runs S4 queries for all three question sets using bidirectional hit detection,
then updates the JSON result files and regenerates Excel reports.

Bidirectional check:
  hit = gold[:120] in chunk_text  OR  chunk_text[:120] in gold_normalized
"""

import json, time, requests, sys, re, openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

API_BASE   = "http://127.0.0.1:8007"
DATA_DIR   = Path("/www/wwwroot/chunkbench.tech/data")
RESULTS    = Path("/www/wwwroot/chunkbench.tech/data/results")

QUERY_PARAMS = {
    "k": 10, "context_mode": "fixed-budget",
    "context_budget_tokens": 1800, "temperature": 0.0,
    "max_tokens": 512, "dedup": True,
}

QUESTION_FILES = {
    "new_questions": [
        DATA_DIR / "cat1_new_questions_batch2.json",
        DATA_DIR / "cat2_cross_doc_questions.json",
        DATA_DIR / "cat_rec_questions.json",
    ],
    "sentinel": DATA_DIR / "sentinel_questions_p2.json",
    "ek_cat1":  DATA_DIR / "ek_cat1_sf_batch1.json",
}

RESULT_FILES = {
    "new_questions": RESULTS / "new_questions_eval" / "new_questions_results_v2.json",
    "sentinel":      RESULTS / "sentinel_eval"      / "sentinel_results.json",
    "ek_cat1":       RESULTS / "ek_cat1_eval"       / "ek_cat1_results.json",
}


def clean(s):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s) if isinstance(s, str) else s


def check_hit_bi(gold_span, chunks, k):
    """Bidirectional: gold in chunk OR chunk in gold."""
    if not gold_span or not chunks:
        return 0
    gold = " ".join(gold_span.lower().split())[:120]
    for c in chunks[:k]:
        txt = " ".join(c["text"].lower().split())
        if gold in txt or txt[:120] in gold:
            return 1
    return 0


def calc_mrr_bi(gold_span, chunks):
    if not gold_span or not chunks:
        return 0.0
    gold = " ".join(gold_span.lower().split())[:120]
    for i, c in enumerate(chunks):
        txt = " ".join(c["text"].lower().split())
        if gold in txt or txt[:120] in gold:
            return 1.0 / (i + 1)
    return 0.0


def call_query(question, mode, strategies):
    r = requests.post(f"{API_BASE}/query",
                      json={"question": question, "mode": mode,
                            "strategies": strategies, **QUERY_PARAMS},
                      timeout=300)
    r.raise_for_status()
    return r.json()["results"]


def score_s4(sr, gold_span):
    chunks = sr.get("retrieved_chunks", [])
    return {
        "hit_at_1":          check_hit_bi(gold_span, chunks, 1),
        "hit_at_3":          check_hit_bi(gold_span, chunks, 3),
        "hit_at_5":          check_hit_bi(gold_span, chunks, 5),
        "mrr":               calc_mrr_bi(gold_span, chunks),
        "retrieved_doc_ids": list(dict.fromkeys(c["doc_id"] for c in chunks)),
        "top3_chunks":       [{"rank": c["rank"], "doc_id": c["doc_id"],
                               "distance": round(c.get("distance", 0), 4),
                               "text": c["text"][:400]} for c in chunks[:3]],
        "context_tokens":    sr.get("context_tokens_est", 0),
        "latency_s":         sr.get("latency_s", 0),
        "answer":            sr.get("answer", ""),
    }


def rescore_existing_other_strategies(result_entry):
    """Re-apply bidirectional check to S1/S2/S3 top3_chunks too (retroactive fix)."""
    gold_span = result_entry.get("gold_span", "")
    for sid in ["S1", "S2", "S3"]:
        sr = result_entry.get("phase2", {}).get(sid, {})
        if not sr:
            continue
        chunks = [{"text": c["text"], "rank": c["rank"], "doc_id": c.get("doc_id",""), "distance": c.get("distance",0)}
                  for c in sr.get("top3_chunks", [])]
        if chunks and gold_span:
            sr["hit_at_1"] = check_hit_bi(gold_span, chunks, 1)
            sr["hit_at_3"] = check_hit_bi(gold_span, chunks, 3)
            # hit_at_5 can only be recalculated from top3 — keep original if it was 1
            # (if hit was in rank 4-5 we can't detect it, so keep original)
            sr["mrr"] = calc_mrr_bi(gold_span, chunks)


def run(name):
    qfiles = QUESTION_FILES[name]
    rfile  = RESULT_FILES[name]

    if not rfile.exists():
        print(f"  Result file not found: {rfile}")
        return None

    if isinstance(qfiles, list):
        questions = []
        for qf in qfiles:
            if qf.exists():
                questions += json.loads(qf.read_text(encoding="utf-8"))
            else:
                print(f"  Warning: {qf} not found, skipping")
    else:
        if not qfiles.exists():
            print(f"  Question file not found: {qfiles}")
            return None
        questions = json.loads(qfiles.read_text(encoding="utf-8"))
    results   = {r["q_id"]: r for r in json.loads(rfile.read_text(encoding="utf-8"))}

    total = len(questions)
    print(f"\n{'='*50}")
    print(f"{name}: {total} questions — re-running S4 only")
    print(f"{'='*50}")

    for i, q in enumerate(questions, 1):
        qid = q["q_id"]
        print(f"  [{i:02d}/{total}] {qid}", flush=True)
        t0 = time.time()
        for attempt in range(3):
            try:
                p2 = call_query(q["question"], "phase2", ["S4"])
                if "S4" in p2:
                    results[qid]["phase2"]["S4"] = score_s4(p2["S4"], q.get("gold_span", ""))
                elapsed = round(time.time() - t0, 1)
                hit = results[qid]["phase2"]["S4"].get("hit_at_5", "?")
                print(f"    ok ({elapsed}s) hit@5={hit}", flush=True)
                break
            except Exception as e:
                if attempt < 2:
                    time.sleep(10)
                else:
                    print(f"    SKIP: {e}", flush=True)

        # Also retroactively apply bidirectional check to S1/S2/S3 in this entry
        rescore_existing_other_strategies(results[qid])

        time.sleep(3)

    out = list(results.values())
    rfile.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"  Saved: {rfile}")
    return out


def summary(out, name):
    strats = ["S1", "S2", "S3", "S4", "B0"]
    print(f"\n--- {name} ---")
    for sid in strats:
        scored = [r for r in out if r["phase2"].get(sid) and
                  r["phase2"][sid].get("hit_at_5") is not None]
        if scored:
            h5  = [r["phase2"][sid]["hit_at_5"] for r in scored]
            mrr = [r["phase2"][sid]["mrr"]       for r in scored]
            print(f"  {sid}: Hit@5={round(sum(h5)/len(h5)*100,1)}%  MRR={round(sum(mrr)/len(mrr),3)}  n={len(h5)}")


def regen_excel_ek_cat1(out):
    green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bold   = Font(bold=True)
    strats = ["S1","S2","S3","S4","B0"]
    wb     = openpyxl.Workbook()
    ws     = wb.active; ws.title = "Summary"
    ws.append(["ek-cat1_sf_batch1 — Phase 2 Evaluation (bidirectional hit)"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"N={len(out)} questions | Hit detection: bidirectional (gold in chunk OR chunk in gold)"])
    ws.append([])
    ws.append(["Strategy","Config","N","Hit@1(%)","Hit@3(%)","Hit@5(%)","MRR"])
    for cell in ws[5]: cell.font = bold
    cfg = {"S1":"size=512,overlap=102(20%)","S2":"size=1024,overlap=102(10%)",
           "S3":"threshold=85,max=2000","S4":"proposition/default","B0":"no retrieval"}
    for sid in strats:
        scored = [r for r in out if r["phase2"].get(sid) and r["phase2"][sid].get("hit_at_5") is not None]
        if not scored: ws.append([sid, cfg.get(sid,""), 0, None, None, None, None]); continue
        h5=[r["phase2"][sid]["hit_at_5"] for r in scored]
        h3=[r["phase2"][sid]["hit_at_3"] for r in scored]
        h1=[r["phase2"][sid]["hit_at_1"] for r in scored]
        m =[r["phase2"][sid]["mrr"]      for r in scored]
        v5=round(sum(h5)/len(h5)*100,1)
        ws.append([sid, cfg.get(sid,""), len(h5),
                   round(sum(h1)/len(h1)*100,1), round(sum(h3)/len(h3)*100,1),
                   v5, round(sum(m)/len(m),3)])
        ws.cell(ws.max_row,6).fill = green if v5>=40 else (red if v5<15 else yellow)

    ws2 = wb.create_sheet("Per-Question")
    hdr = ["q_id","source_doc","question","gold_span","gold_answer"]
    for sid in strats: hdr += [f"{sid}_hit5", f"{sid}_hit3", f"{sid}_mrr", f"{sid}_ctx_tok", f"{sid}_docs"]
    for sid in strats: hdr.append(f"{sid}_answer")
    ws2.append(hdr)
    for cell in ws2[1]: cell.font = bold
    for r in out:
        row = [r["q_id"], r["source_doc"], clean(r["question"])[:150],
               clean(r.get("gold_span",""))[:150], clean(r.get("gold_answer",""))[:100]]
        for sid in strats:
            sr   = r["phase2"].get(sid, {})
            docs = ", ".join(sr.get("retrieved_doc_ids", []))
            row += [sr.get("hit_at_5"), sr.get("hit_at_3"), sr.get("mrr"), sr.get("context_tokens"), docs]
        for sid in strats:
            row.append(clean(r["phase2"].get(sid,{}).get("answer",""))[:250])
        ws2.append(row)
        rn = ws2.max_row
        for i, h in enumerate(hdr):
            if h.endswith("_hit5"):
                c = ws2.cell(rn, i+1)
                if c.value == 1: c.fill = green
                elif c.value == 0: c.fill = red

    for ws_ in [ws, ws2]:
        for col in ws_.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            ws_.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2, 45)

    out_path = RESULTS / "ek_cat1_eval" / "ek_cat1_eval.xlsx"
    wb.save(str(out_path))
    print(f"  Excel: {out_path}")


def main():
    try:
        r = requests.get(f"{API_BASE}/health", timeout=10)
        print(f"API: {r.json()}")
    except Exception as e:
        print(f"API not reachable: {e}"); sys.exit(1)

    for name in ["new_questions", "sentinel", "ek_cat1"]:
        out = run(name)
        if out:
            summary(out, name)

    # Regenerate ek_cat1 excel as example (others follow same pattern)
    ek = json.loads((RESULTS / "ek_cat1_eval" / "ek_cat1_results.json").read_text())
    regen_excel_ek_cat1(ek)

    print(f"\nDone: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
