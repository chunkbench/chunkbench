#!/usr/bin/env python3
"""
Evaluation script for new question batches (cat1_b2, cat2_cross_doc, cat_rec).
Uses /query endpoint (returns full retrieved_chunks) for accurate scoring.
Phase 1: all 5 strategies with default indexes.
Phase 2: S1/S2/S3 with best-config indexes; S4+B0 copied from Phase 1.

Usage:
    python run_new_questions_eval.py [--resume]
"""

import json, time, requests, sys, argparse
from pathlib import Path
from datetime import datetime

API_BASE    = "http://127.0.0.1:8007"
DATA_DIR    = Path("/www/wwwroot/chunkbench.tech/data")
RESULTS_DIR = Path("/www/wwwroot/chunkbench.tech/data/results/new_questions_eval")
CHECKPOINT  = RESULTS_DIR / "checkpoint_v2.json"

QUESTION_FILES = {
    "cat1_b2": DATA_DIR / "cat1_new_questions_batch2.json",
    "cat2":    DATA_DIR / "cat2_cross_doc_questions.json",
    "cat_rec": DATA_DIR / "cat_rec_questions.json",
}

SLEEP_BETWEEN_PHASE = 5   # seconds between phase1 and phase2 calls per question
SLEEP_BETWEEN_Q     = 4   # extra sleep between questions

PHASE1_STRATEGIES   = ["S1", "S2", "S3", "S4", "B0"]
PHASE2_NEW_STRATS   = ["S1", "S2", "S3"]   # S4 and B0 unchanged

QUERY_PARAMS = {
    "k": 10,
    "context_mode": "fixed-budget",
    "context_budget_tokens": 1800,
    "temperature": 0.0,
    "max_tokens": 512,
    "dedup": True,
}

RESULTS_DIR.mkdir(parents=True, exist_ok=True)


def load_questions():
    questions = []
    for group, path in QUESTION_FILES.items():
        if not path.exists():
            print(f"  WARNING: {path} not found, skipping.")
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        for q in data:
            entry = {
                "group":       group,
                "q_id":        q["q_id"],
                "question":    q["question"],
                "category":    q.get("category", group),
                "gold_answer": q.get("gold_answer", ""),
                "gold_span":   q.get("gold_span") or q.get("gold_span_a", ""),
                "gold_span_2": q.get("gold_span_2") or q.get("gold_span_b", ""),
                "source_doc":  q.get("source_doc") or q.get("source_doc_a", ""),
                "source_doc_b": q.get("source_doc_b", ""),
            }
            questions.append(entry)
    return questions


def _norm(d):
    if not d: return d
    if d.endswith(".pdf"): d = d[:-4]
    return d.replace("_shallow", "").replace("_deep", "")

def check_hit(source_doc, chunks, k):
    if not source_doc or not chunks:
        return 0
    src = _norm(source_doc)
    return int(any(_norm(c.get("doc_id","")) == src for c in chunks[:k]))


def reciprocal_rank(source_doc, chunks):
    if not source_doc or not chunks:
        return 0.0
    src = _norm(source_doc)
    for i, c in enumerate(chunks):
        if _norm(c.get("doc_id","")) == src:
            return 1.0 / (i + 1)
    return 0.0


def score_strategy(strat_result, source_doc, source_doc_b=None):
    chunks = strat_result.get("retrieved_chunks", [])
    out = {
        "hit_at_1":          check_hit(source_doc, chunks, 1),
        "hit_at_3":          check_hit(source_doc, chunks, 3),
        "hit_at_5":          check_hit(source_doc, chunks, 5),
        "mrr":               reciprocal_rank(source_doc, chunks),
        "retrieved_doc_ids": list(dict.fromkeys(c["doc_id"] for c in chunks)),
        "top3_chunks": [
            {
                "rank":     c["rank"],
                "doc_id":   c["doc_id"],
                "distance": round(c.get("distance", 0), 4),
                "text":     c["text"][:400],
            }
            for c in chunks[:3]
        ],
        "context_tokens": strat_result.get("context_tokens_est", 0),
        "latency_s":      strat_result.get("latency_s", 0),
        "answer":         strat_result.get("answer", ""),
        "prompt_tokens":  strat_result.get("prompt_tokens", 0),
    }
    if source_doc_b:
        out["hit_at_5_doc_b"] = check_hit(source_doc_b, chunks, 5)
        out["mrr_doc_b"]      = reciprocal_rank(source_doc_b, chunks)
        out["both_hit_at_5"]  = int(out["hit_at_5"] == 1 and out["hit_at_5_span2"] == 1)
    return out


def call_query(question, mode, strategies):
    payload = {
        "question":   question,
        "mode":       mode,
        "strategies": strategies,
        **QUERY_PARAMS,
    }
    r = requests.post(f"{API_BASE}/query", json=payload, timeout=300)
    r.raise_for_status()
    return r.json()["results"]


def run_eval(resume=False):
    questions = load_questions()
    print(f"\nLoaded {len(questions)} questions:")
    for g in QUESTION_FILES:
        n = sum(1 for q in questions if q["group"] == g)
        print(f"  {g}: {n}")

    done_ids = set()
    results  = {}
    if resume and CHECKPOINT.exists():
        cp = json.loads(CHECKPOINT.read_text())
        done_ids = set(cp.get("done_ids", []))
        results  = cp.get("results", {})
        print(f"\nResuming: {len(done_ids)} done, {len(questions)-len(done_ids)} remaining.")

    remaining = [q for q in questions if q["q_id"] not in done_ids]
    total     = len(questions)

    for q in remaining:
        n = len(done_ids) + 1
        print(f"\n[{n:02d}/{total}] {q['q_id']} ({q['group']})", flush=True)
        print(f"  Q: {q['question'][:80]}...", flush=True)
        t_q = time.time()

        result = {
            "q_id":        q["q_id"],
            "group":       q["group"],
            "question":    q["question"],
            "category":    q["category"],
            "gold_answer": q["gold_answer"],
            "gold_span":   q["gold_span"],
            "gold_span_2": q["gold_span_2"],
            "source_doc":  q["source_doc"],
            "source_doc_b": q["source_doc_b"],
            "phase1": {},
            "phase2": {},
        }

        # ── Phase 1: all 5 strategies ──────────────────────────
        print(f"  [P1] querying {PHASE1_STRATEGIES}...", end=" ", flush=True)
        for attempt in range(3):
            try:
                p1 = call_query(q["question"], mode="phase1", strategies=PHASE1_STRATEGIES)
                for sid in PHASE1_STRATEGIES:
                    if sid in p1:
                        result["phase1"][sid] = score_strategy(
                            p1[sid], q["source_doc"], q.get("source_doc_b")
                        )
                elapsed = round(time.time() - t_q, 1)
                print(f"ok ({elapsed}s)", flush=True)
                break
            except Exception as e:
                print(f"\n  attempt {attempt+1} failed: {e}", flush=True)
                if attempt < 2:
                    time.sleep(15)
                else:
                    print("  SKIP Phase 1.", flush=True)

        time.sleep(SLEEP_BETWEEN_PHASE)

        # ── Phase 2: S1/S2/S3 with best-config indexes ─────────
        print(f"  [P2] querying {PHASE2_NEW_STRATS}...", end=" ", flush=True)
        for attempt in range(3):
            try:
                p2 = call_query(q["question"], mode="phase2", strategies=PHASE2_NEW_STRATS)
                for sid in PHASE2_NEW_STRATS:
                    if sid in p2:
                        result["phase2"][sid] = score_strategy(
                            p2[sid], q["source_doc"], q.get("source_doc_b")
                        )
                # S4 and B0 are identical in phase2 — copy from phase1
                for sid in ["S4", "B0"]:
                    if sid in result["phase1"]:
                        result["phase2"][sid] = result["phase1"][sid]
                elapsed = round(time.time() - t_q, 1)
                print(f"ok ({elapsed}s total)", flush=True)
                break
            except Exception as e:
                print(f"\n  attempt {attempt+1} failed: {e}", flush=True)
                if attempt < 2:
                    time.sleep(15)
                else:
                    print("  SKIP Phase 2.", flush=True)

        # Hit summary line
        p1_hits = " ".join(
            f"{s}:{'Y' if result['phase1'].get(s,{}).get('hit_at_5')==1 else 'N'}"
            for s in ["S1","S2","S3","S4"]
        )
        p2_hits = " ".join(
            f"{s}:{'Y' if result['phase2'].get(s,{}).get('hit_at_5')==1 else 'N'}"
            for s in ["S1","S2","S3","S4"]
        )
        print(f"  P1: {p1_hits}  |  P2: {p2_hits}", flush=True)

        results[q["q_id"]] = result
        done_ids.add(q["q_id"])
        CHECKPOINT.write_text(
            json.dumps({"done_ids": list(done_ids), "results": results}, ensure_ascii=False)
        )
        time.sleep(SLEEP_BETWEEN_Q)

    print(f"\nAll {total} questions done.", flush=True)

    out_json = RESULTS_DIR / "new_questions_results_v2.json"
    out_json.write_text(json.dumps(list(results.values()), indent=2, ensure_ascii=False))
    print(f"JSON: {out_json}", flush=True)
    return list(results.values())


def aggregate(results, phase, strategies):
    agg = {}
    for sid in strategies:
        scored = [r for r in results if r[phase].get(sid) and r[phase][sid].get("hit_at_5") is not None]
        if not scored:
            agg[sid] = {"n": 0, "hit_at_5": None, "hit_at_3": None, "hit_at_1": None, "mrr": None}
            continue
        h5  = [r[phase][sid]["hit_at_5"] for r in scored]
        h3  = [r[phase][sid]["hit_at_3"] for r in scored]
        h1  = [r[phase][sid]["hit_at_1"] for r in scored]
        mrr = [r[phase][sid]["mrr"]      for r in scored]
        agg[sid] = {
            "n":       len(h5),
            "hit_at_5": round(sum(h5)/len(h5)*100, 1),
            "hit_at_3": round(sum(h3)/len(h3)*100, 1),
            "hit_at_1": round(sum(h1)/len(h1)*100, 1),
            "mrr":      round(sum(mrr)/len(mrr), 3),
        }
    return agg


def generate_excel(results):
    import openpyxl, re
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    def clean(s):
        if not isinstance(s, str): return s
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)

    green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bold   = Font(bold=True)

    strats = ["S1", "S2", "S3", "S4", "B0"]
    groups = list(QUESTION_FILES.keys())
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["New Questions Evaluation — Phase 1 vs Phase 2"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"Total questions: {len(results)}  |  Scoring: doc-level (source_doc in top-k retrieved doc_ids)"])
    ws.append([])

    for phase in ["phase1", "phase2"]:
        label = "PHASE 1 — default configs" if phase == "phase1" else "PHASE 2 — best configs"
        phase_configs = {
            "S1": "size=512, overlap=50 (10%)" if phase == "phase1" else "size=512, overlap=102 (20%)",
            "S2": "size=512, overlap=50 (10%)" if phase == "phase1" else "size=1024, overlap=102 (10%)",
            "S3": "percentile/95"              if phase == "phase1" else "threshold=85, max=2000",
            "S4": "proposition/default (same both phases)",
            "B0": "no retrieval (same both phases)",
        }
        ws.append([label])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        ws.append(["Group", "Strategy", "Config", "N", "Hit@1(%)", "Hit@3(%)", "Hit@5(%)", "MRR"])
        for cell in ws[ws.max_row]: cell.font = bold

        for grp in groups + ["ALL"]:
            grp_res = [r for r in results if r["group"] == grp] if grp != "ALL" else results
            if not grp_res: continue
            agg = aggregate(grp_res, phase, strats)
            for sid in strats:
                a = agg[sid]
                ws.append([grp, sid, phase_configs[sid], a["n"],
                            a["hit_at_1"], a["hit_at_3"], a["hit_at_5"], a["mrr"]])
                h5_cell = ws.cell(ws.max_row, 7)
                if a["hit_at_5"] is not None:
                    h5_cell.fill = green if a["hit_at_5"] >= 30 else (red if a["hit_at_5"] < 10 else yellow)
        ws.append([])

    # ── Sheet 2: Phase Gain ───────────────────────────────────────
    ws2 = wb.create_sheet("P1 vs P2 Gain")
    ws2.append(["Strategy", "Group", "N",
                "P1 Hit@5(%)", "P2 Hit@5(%)", "Δ Hit@5",
                "P1 MRR",     "P2 MRR",      "Δ MRR",
                "P1 config", "P2 config"])
    for cell in ws2[1]: cell.font = bold

    cfg = {
        "S1": ("size=512,ov=50(10%)", "size=512,ov=102(20%)"),
        "S2": ("size=512,ov=50(10%)", "size=1024,ov=102(10%)"),
        "S3": ("percentile/95",       "threshold=85,max=2000"),
    }
    for grp in groups + ["ALL"]:
        grp_res = [r for r in results if r["group"] == grp] if grp != "ALL" else results
        if not grp_res: continue
        a1 = aggregate(grp_res, "phase1", strats)
        a2 = aggregate(grp_res, "phase2", strats)
        for sid in ["S1","S2","S3"]:
            p1h = a1[sid]["hit_at_5"]; p2h = a2[sid]["hit_at_5"]
            p1m = a1[sid]["mrr"];      p2m = a2[sid]["mrr"]
            dh = round(p2h-p1h,1) if p1h is not None and p2h is not None else None
            dm = round(p2m-p1m,3) if p1m is not None and p2m is not None else None
            ws2.append([sid, grp, a1[sid]["n"], p1h, p2h, dh, p1m, p2m, dm, cfg[sid][0], cfg[sid][1]])
            if dh is not None:
                ws2.cell(ws2.max_row, 6).fill = green if dh > 0 else (red if dh < 0 else yellow)

    # ── Sheet 3: Per-Question Detail ──────────────────────────────
    ws3 = wb.create_sheet("Per-Question")
    hdr = ["q_id","group","category","source_doc","source_doc_b","question","gold_span","gold_span_2","gold_answer"]
    for ph in ["P1","P2"]:
        for sid in strats:
            hdr += [f"{ph}_{sid}_hit5", f"{ph}_{sid}_hit3", f"{ph}_{sid}_mrr",
                    f"{ph}_{sid}_ctx_tok", f"{ph}_{sid}_latency",
                    f"{ph}_{sid}_retrieved_docs"]
    for ph in ["P1","P2"]:
        for sid in strats:
            hdr.append(f"{ph}_{sid}_answer")
    ws3.append(hdr)
    for cell in ws3[1]: cell.font = bold

    ph_map = {"P1": "phase1", "P2": "phase2"}
    for r in results:
        row = [r["q_id"], r["group"], r["category"],
               r["source_doc"], r.get("source_doc_b",""),
               clean(r["question"])[:150],
               clean(r["gold_span"])[:150],
               clean(r.get("gold_span_2",""))[:100],
               clean(r["gold_answer"])[:100]]
        for ph_lbl, ph_key in ph_map.items():
            for sid in strats:
                sr = r[ph_key].get(sid, {})
                docs = ", ".join(sr.get("retrieved_doc_ids", []))
                row += [sr.get("hit_at_5"), sr.get("hit_at_3"), sr.get("mrr"),
                        sr.get("context_tokens"), sr.get("latency_s"), docs]
        for ph_lbl, ph_key in ph_map.items():
            for sid in strats:
                sr = r[ph_key].get(sid, {})
                row.append(clean(sr.get("answer",""))[:250])
        ws3.append(row)
        rn = ws3.max_row
        hit5_cols = [i+1 for i,h in enumerate(hdr) if h.endswith("_hit5")]
        for col in hit5_cols:
            c = ws3.cell(rn, col)
            if c.value == 1: c.fill = green
            elif c.value == 0: c.fill = red

    # ── Sheet 4: Retrieved Documents ─────────────────────────────
    ws4 = wb.create_sheet("Retrieved Docs")
    ws4.append(["q_id","group","gold_source_doc","phase","strategy",
                "rank","retrieved_doc_id","distance","is_gold_doc","chunk_text_preview"])
    for cell in ws4[1]: cell.font = bold

    for r in results:
        gold_docs = set(filter(None, [r["source_doc"], r.get("source_doc_b","")]))
        for ph_key in ["phase1","phase2"]:
            for sid in strats:
                sr = r[ph_key].get(sid, {})
                for chunk in sr.get("top3_chunks", []):
                    is_gold = "YES" if chunk["doc_id"] in gold_docs else ""
                    ws4.append([r["q_id"], r["group"], r["source_doc"],
                                ph_key, sid, chunk["rank"], chunk["doc_id"],
                                chunk["distance"], is_gold, clean(chunk["text"])[:250]])
                    if is_gold:
                        ws4.cell(ws4.max_row, 9).fill = green

    for ws_ in [ws, ws2, ws3, ws4]:
        for col in ws_.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            ws_.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2, 45)

    out = RESULTS_DIR / "new_questions_eval_v2.xlsx"
    wb.save(str(out))
    print(f"Excel: {out}", flush=True)


def print_summary(results):
    print("\n" + "="*60)
    strats = ["S1","S2","S3","S4","B0"]
    for phase in ["phase1","phase2"]:
        label = "Phase 1" if phase == "phase1" else "Phase 2"
        print(f"\n{label}:")
        agg = aggregate(results, phase, strats)
        for sid in strats:
            a = agg[sid]
            if a["hit_at_5"] is not None:
                print(f"  {sid}: Hit@5={a['hit_at_5']}%  MRR={a['mrr']}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true")
    args = parser.parse_args()

    try:
        r = requests.get(f"{API_BASE}/health", timeout=10)
        print(f"API: {r.json()}")
    except Exception as e:
        print(f"API not reachable: {e}"); sys.exit(1)

    qs = load_questions()
    print(f"\nStarted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    est = len(qs) * 160 // 60
    print(f"Estimated: ~{est} min ({len(qs)} questions × ~160s each)")

    results = run_eval(resume=args.resume)
    print_summary(results)
    generate_excel(results)
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
