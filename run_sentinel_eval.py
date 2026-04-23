#!/usr/bin/env python3
"""
Sentinel questions evaluation — Phase 1 and Phase 2.
Usage: python run_sentinel_eval.py [--resume]
"""

import json, time, requests, sys, argparse, re, openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

API_BASE    = "http://127.0.0.1:8007"
DATA_DIR    = Path("/www/wwwroot/chunkbench.tech/data")
RESULTS_DIR = Path("/www/wwwroot/chunkbench.tech/data/results/sentinel_eval")
CHECKPOINT  = RESULTS_DIR / "checkpoint.json"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

QUESTION_FILE = DATA_DIR / "sentinel_questions_p2.json"

QUERY_PARAMS = {
    "k": 10, "context_mode": "fixed-budget",
    "context_budget_tokens": 1800, "temperature": 0.0,
    "max_tokens": 512, "dedup": True,
}
P1_STRATS = ["S1", "S2", "S3", "S4", "B0"]
P2_STRATS = ["S1", "S2", "S3"]


def clean(s):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s) if isinstance(s, str) else s


def check_hit(gold, chunks, k):
    if not gold or not chunks: return 0
    g = " ".join(gold.lower().split())[:120]
    return int(any(g in " ".join(c["text"].lower().split()) for c in chunks[:k]))


def calc_mrr(gold, chunks):
    if not gold or not chunks: return 0.0
    g = " ".join(gold.lower().split())[:120]
    for i, c in enumerate(chunks):
        if g in " ".join(c["text"].lower().split()):
            return 1.0 / (i + 1)
    return 0.0


def score(sr, gold_span):
    chunks = sr.get("retrieved_chunks", [])
    return {
        "hit_at_1":          check_hit(gold_span, chunks, 1),
        "hit_at_3":          check_hit(gold_span, chunks, 3),
        "hit_at_5":          check_hit(gold_span, chunks, 5),
        "mrr":               calc_mrr(gold_span, chunks),
        "retrieved_doc_ids": list(dict.fromkeys(c["doc_id"] for c in chunks)),
        "top3_chunks": [
            {"rank": c["rank"], "doc_id": c["doc_id"],
             "distance": round(c.get("distance", 0), 4), "text": c["text"][:400]}
            for c in chunks[:3]
        ],
        "context_tokens": sr.get("context_tokens_est", 0),
        "latency_s":      sr.get("latency_s", 0),
        "answer":         sr.get("answer", ""),
    }


def call_query(question, mode, strategies):
    r = requests.post(
        f"{API_BASE}/query",
        json={"question": question, "mode": mode, "strategies": strategies, **QUERY_PARAMS},
        timeout=300,
    )
    r.raise_for_status()
    return r.json()["results"]


def run_eval(resume=False):
    questions = json.loads(QUESTION_FILE.read_text(encoding="utf-8"))
    total = len(questions)

    done_ids, results = set(), {}
    if resume and CHECKPOINT.exists():
        cp = json.loads(CHECKPOINT.read_text())
        done_ids = set(cp["done_ids"])
        results  = cp["results"]
        print(f"Resuming: {len(done_ids)} done, {total-len(done_ids)} remaining.")

    remaining = [q for q in questions if q["q_id"] not in done_ids]
    print(f"Loaded {total} sentinel questions, {len(remaining)} to run.", flush=True)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)

    for q in remaining:
        n = len(done_ids) + 1
        print(f"\n[{n:02d}/{total}] {q['q_id']} ({q.get('needle_type','')}) — {q['question'][:70]}...", flush=True)
        t0 = time.time()
        res = {
            "q_id":        q["q_id"],
            "category":    q.get("category", "sentinel"),
            "needle_type": q.get("needle_type", ""),
            "question":    q["question"],
            "gold_answer": q.get("gold_answer", ""),
            "gold_span":   q.get("gold_span", ""),
            "source_doc":  q.get("source_doc", ""),
            "phase1": {}, "phase2": {},
        }

        # Phase 1
        print(f"  [P1]...", end=" ", flush=True)
        for attempt in range(3):
            try:
                p1 = call_query(q["question"], "phase1", P1_STRATS)
                for sid in P1_STRATS:
                    if sid in p1:
                        res["phase1"][sid] = score(p1[sid], q["gold_span"])
                print(f"ok ({round(time.time()-t0, 1)}s)", flush=True)
                break
            except Exception as e:
                if attempt < 2: time.sleep(15)
                else: print(f"SKIP: {e}", flush=True)

        time.sleep(5)

        # Phase 2
        print(f"  [P2]...", end=" ", flush=True)
        for attempt in range(3):
            try:
                p2 = call_query(q["question"], "phase2", P2_STRATS)
                for sid in P2_STRATS:
                    if sid in p2:
                        res["phase2"][sid] = score(p2[sid], q["gold_span"])
                for sid in ["S4", "B0"]:
                    if sid in res["phase1"]:
                        res["phase2"][sid] = res["phase1"][sid]
                print(f"ok ({round(time.time()-t0, 1)}s total)", flush=True)
                break
            except Exception as e:
                if attempt < 2: time.sleep(15)
                else: print(f"SKIP: {e}", flush=True)

        p1h = " ".join(f"{s}:{'Y' if res['phase1'].get(s,{}).get('hit_at_5')==1 else 'N'}"
                       for s in ["S1","S2","S3","S4"])
        p2h = " ".join(f"{s}:{'Y' if res['phase2'].get(s,{}).get('hit_at_5')==1 else 'N'}"
                       for s in ["S1","S2","S3","S4"])
        print(f"  P1: {p1h}  |  P2: {p2h}", flush=True)

        results[q["q_id"]] = res
        done_ids.add(q["q_id"])
        CHECKPOINT.write_text(
            json.dumps({"done_ids": list(done_ids), "results": results}, ensure_ascii=False)
        )
        time.sleep(4)

    out = list(results.values())
    (RESULTS_DIR / "sentinel_results.json").write_text(
        json.dumps(out, indent=2, ensure_ascii=False)
    )
    print(f"\nJSON saved.", flush=True)
    return out


def generate_excel(out):
    green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bold   = Font(bold=True)
    strats = ["S1", "S2", "S3", "S4", "B0"]
    wb     = openpyxl.Workbook()

    # ── Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Sentinel Questions Evaluation — Phase 1 vs Phase 2"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"N={len(out)} sentinel questions"])
    ws.append([])

    for phase, label in [("phase1", "PHASE 1 — default configs"),
                          ("phase2", "PHASE 2 — best configs")]:
        ws.append([label])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        ws.append(["Strategy", "N", "Hit@1(%)", "Hit@3(%)", "Hit@5(%)", "MRR"])
        for cell in ws[ws.max_row]: cell.font = bold
        for sid in strats:
            scored = [r for r in out if r[phase].get(sid) and
                      r[phase][sid].get("hit_at_5") is not None]
            if not scored:
                ws.append([sid, 0, None, None, None, None])
                continue
            h5  = [r[phase][sid]["hit_at_5"] for r in scored]
            h3  = [r[phase][sid]["hit_at_3"] for r in scored]
            h1  = [r[phase][sid]["hit_at_1"] for r in scored]
            m   = [r[phase][sid]["mrr"]      for r in scored]
            v5  = round(sum(h5)/len(h5)*100, 1)
            ws.append([sid, len(h5),
                       round(sum(h1)/len(h1)*100, 1),
                       round(sum(h3)/len(h3)*100, 1),
                       v5, round(sum(m)/len(m), 3)])
            ws.cell(ws.max_row, 5).fill = (
                green if v5 >= 50 else (red if v5 < 20 else yellow)
            )
        ws.append([])

    # Phase gain row
    ws.append(["Phase 2 vs Phase 1 Gain (S1/S2/S3)"])
    ws.cell(ws.max_row, 1).font = bold
    ws.append(["Strategy", "P1 Hit@5(%)", "P2 Hit@5(%)", "Δ Hit@5", "P1 MRR", "P2 MRR", "Δ MRR"])
    for cell in ws[ws.max_row]: cell.font = bold
    for sid in ["S1", "S2", "S3"]:
        def agg(phase):
            sc = [r for r in out if r[phase].get(sid) and
                  r[phase][sid].get("hit_at_5") is not None]
            if not sc: return None, None
            h5 = [r[phase][sid]["hit_at_5"] for r in sc]
            m  = [r[phase][sid]["mrr"]      for r in sc]
            return round(sum(h5)/len(h5)*100, 1), round(sum(m)/len(m), 3)
        p1h, p1m = agg("phase1")
        p2h, p2m = agg("phase2")
        dh = round(p2h-p1h, 1) if p1h is not None and p2h is not None else None
        dm = round(p2m-p1m, 3) if p1m is not None and p2m is not None else None
        ws.append([sid, p1h, p2h, dh, p1m, p2m, dm])
        if dh is not None:
            ws.cell(ws.max_row, 4).fill = (
                green if dh > 0 else (red if dh < 0 else yellow)
            )

    # ── Per-Question ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Question")
    hdr = ["q_id", "needle_type", "source_doc", "question", "gold_span", "gold_answer"]
    for ph in ["P1", "P2"]:
        for sid in strats:
            hdr += [f"{ph}_{sid}_hit5", f"{ph}_{sid}_mrr",
                    f"{ph}_{sid}_ctx_tok", f"{ph}_{sid}_docs"]
    for ph in ["P1", "P2"]:
        for sid in strats:
            hdr.append(f"{ph}_{sid}_answer")
    ws2.append(hdr)
    for cell in ws2[1]: cell.font = bold

    for r in out:
        row = [r["q_id"], r["needle_type"], r["source_doc"],
               clean(r["question"])[:150], clean(r["gold_span"])[:150],
               clean(r["gold_answer"])[:100]]
        for ph_lbl, ph_key in [("P1","phase1"), ("P2","phase2")]:
            for sid in strats:
                sr   = r[ph_key].get(sid, {})
                docs = ", ".join(sr.get("retrieved_doc_ids", []))
                row += [sr.get("hit_at_5"), sr.get("mrr"),
                        sr.get("context_tokens"), docs]
        for ph_lbl, ph_key in [("P1","phase1"), ("P2","phase2")]:
            for sid in strats:
                row.append(clean(r[ph_key].get(sid, {}).get("answer", ""))[:200])
        ws2.append(row)
        rn = ws2.max_row
        for i, h in enumerate(hdr):
            if h.endswith("_hit5"):
                c = ws2.cell(rn, i+1)
                if c.value == 1:   c.fill = green
                elif c.value == 0: c.fill = red

    # ── Retrieved Docs ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Retrieved Docs")
    ws3.append(["q_id", "needle_type", "gold_source", "phase", "strategy",
                "rank", "doc_id", "distance", "is_gold", "text"])
    for cell in ws3[1]: cell.font = bold
    for r in out:
        for ph_key in ["phase1", "phase2"]:
            for sid in strats:
                for chunk in r[ph_key].get(sid, {}).get("top3_chunks", []):
                    is_gold = "YES" if chunk["doc_id"] == r["source_doc"] else ""
                    ws3.append([r["q_id"], r["needle_type"], r["source_doc"],
                                ph_key, sid, chunk["rank"], chunk["doc_id"],
                                chunk["distance"], is_gold, clean(chunk["text"])[:250]])
                    if is_gold: ws3.cell(ws3.max_row, 9).fill = green

    for ws_ in [ws, ws2, ws3]:
        for col in ws_.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            ws_.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2, 45)

    out_path = RESULTS_DIR / "sentinel_eval.xlsx"
    wb.save(str(out_path))
    print(f"Excel: {out_path}", flush=True)


def print_summary(out):
    strats = ["S1", "S2", "S3", "S4", "B0"]
    print("\n" + "="*50, flush=True)
    for phase in ["phase1", "phase2"]:
        print(f"\n{'Phase 1' if phase=='phase1' else 'Phase 2'}:", flush=True)
        for sid in strats:
            scored = [r for r in out if r[phase].get(sid) and
                      r[phase][sid].get("hit_at_5") is not None]
            if scored:
                h5 = [r[phase][sid]["hit_at_5"] for r in scored]
                m  = [r[phase][sid]["mrr"]      for r in scored]
                print(f"  {sid}: Hit@5={round(sum(h5)/len(h5)*100,1)}%  MRR={round(sum(m)/len(m),3)}", flush=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true")
    args = parser.parse_args()

    try:
        r = requests.get(f"{API_BASE}/health", timeout=10)
        print(f"API: {r.json()}", flush=True)
    except Exception as e:
        print(f"API not reachable: {e}"); sys.exit(1)

    out = run_eval(resume=args.resume)
    print_summary(out)
    generate_excel(out)
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)


if __name__ == "__main__":
    main()
