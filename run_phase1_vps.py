#!/usr/bin/env python3
import json, requests, time, os, sys
from datetime import datetime

QUESTIONS_FILE = "data/questions_v1_FROZEN.json"
RESULTS_DIR    = "data/results/phase1"
CHECKPOINT     = f"{RESULTS_DIR}/checkpoint.json"
API_BASE       = "http://127.0.0.1:8007"

STRATEGIES = ["S1", "S2", "S3", "S4", "B0"]
EVAL_PARAMS = {
    "strategies":            STRATEGIES,
    "k":                     10,
    "context_mode":          "fixed-budget",
    "context_budget_tokens": 1800,
    "temperature":           0.0,
    "dedup":                 True,
}

os.makedirs(RESULTS_DIR, exist_ok=True)


def load_checkpoint():
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT) as f:
            c = json.load(f)
        return set(c.get("done_ids", [])), c.get("results", [])
    return set(), []


def save_checkpoint(done_ids, results):
    with open(CHECKPOINT, "w") as f:
        json.dump({"done_ids": list(done_ids), "results": results,
                   "updated": datetime.now().isoformat()}, f, indent=2)


def run_question(q):
    payload = {
        "questions": [{
            "q_id":        q["q_id"],
            "question":    q["question"],
            "category":    q["category"],
            "needle_type": q.get("needle_type", "corpus"),
            "gold_answer": q.get("gold_answer", ""),
            "gold_span":   q.get("gold_span", ""),
            "gold_span_2": q.get("gold_span_2"),
            "source_doc":  q.get("source_doc", ""),
            "difficulty":  q.get("difficulty", "medium"),
        }],
        **EVAL_PARAMS,
    }
    r = requests.post(f"{API_BASE}/evaluate", json=payload, timeout=300)
    r.raise_for_status()
    return r.json()["results"][0]


def generate_excel(all_results, output_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["HAE-RAG Phase 1 Results", "", "", "", ""])
    ws.append(["Config: k=10, fixed-budget, 1800 tokens, dedup=0.95"])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    ws.append(["Strategy", "Hit@5 (%)", "Hit@3 (%)", "Hit@1 (%)", "MRR", "Avg Context Tokens"])
    header_row = ws.max_row
    for sid in STRATEGIES:
        h5 = [r["strategies"][sid]["hit_at_5"] for r in all_results
              if r["strategies"].get(sid) and r["strategies"][sid]["hit_at_5"] is not None]
        h3 = [r["strategies"][sid].get("hit_at_3", 0) for r in all_results
              if r["strategies"].get(sid) and r["strategies"][sid].get("hit_at_3") is not None]
        h1 = [r["strategies"][sid].get("hit_at_1", 0) for r in all_results
              if r["strategies"].get(sid) and r["strategies"][sid].get("hit_at_1") is not None]
        mrr = [r["strategies"][sid]["mrr"] for r in all_results
               if r["strategies"].get(sid) and r["strategies"][sid]["mrr"] is not None]
        ctx = [r["strategies"][sid].get("context_tokens", 0) for r in all_results
               if r["strategies"].get(sid)]
        if h5:
            ws.append([sid,
                       round(sum(h5)/len(h5)*100, 1),
                       round(sum(h3)/len(h3)*100, 1) if h3 else "N/A",
                       round(sum(h1)/len(h1)*100, 1) if h1 else "N/A",
                       round(sum(mrr)/len(mrr), 3) if mrr else "N/A",
                       round(sum(ctx)/len(ctx)) if ctx else 0])
        else:
            ws.append([sid, "N/A", "N/A", "N/A", "N/A", 0])

    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    ws.append([])
    ws.append(["Category Breakdown (Hit@5 %)"])
    cats = ["single_fact", "numerical", "multi_hop", "negation", "sentinel"]
    ws.append(["Category"] + list(["S1", "S2", "S3", "S4"]))
    for cat in cats:
        row = [cat]
        for sid in ["S1", "S2", "S3", "S4"]:
            qs = [r for r in all_results if r["category"] == cat]
            h5 = [r["strategies"][sid]["hit_at_5"] for r in qs
                  if r["strategies"].get(sid) and r["strategies"][sid]["hit_at_5"] is not None]
            row.append(f"{sum(h5)/len(h5)*100:.0f}%" if h5 else "N/A")
        ws.append(row)

    ws2 = wb.create_sheet("Per-Question")
    headers = ["q_id", "question", "category", "needle_type", "difficulty",
               "gold_answer", "source_doc",
               "S1_hit5", "S1_mrr", "S1_answer", "S1_tokens",
               "S2_hit5", "S2_mrr", "S2_answer", "S2_tokens",
               "S3_hit5", "S3_mrr", "S3_answer", "S3_tokens",
               "S4_hit5", "S4_mrr", "S4_answer", "S4_tokens",
               "B0_answer"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for r in all_results:
        row = [
            r.get("q_id", ""),
            r.get("question", "")[:100],
            r.get("category", ""),
            r.get("needle_type", ""),
            r.get("difficulty", ""),
            r.get("gold_answer", ""),
            r.get("source_doc", "") if isinstance(r.get("source_doc"), str) else "",
        ]
        for sid in ["S1", "S2", "S3", "S4"]:
            sr = r["strategies"].get(sid, {})
            row.extend([
                sr.get("hit_at_5", ""),
                round(sr["mrr"], 3) if sr.get("mrr") is not None else "",
                (sr.get("answer", "") or "")[:200],
                sr.get("context_tokens", 0),
            ])
        b0 = r["strategies"].get("B0", {})
        row.append((b0.get("answer", "") or "")[:200])
        ws2.append(row)

        row_num = ws2.max_row
        for col_offset, sid in enumerate(["S1", "S2", "S3", "S4"]):
            col = 8 + col_offset * 4
            cell = ws2.cell(row=row_num, column=col)
            if cell.value == 1:
                cell.fill = green
            elif cell.value == 0:
                cell.fill = red

    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    ws3 = wb.create_sheet("Retrieved Chunks")
    ws3.append(["q_id", "strategy", "rank", "doc_id", "distance", "text_preview"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for r in all_results:
        for sid in ["S1", "S2", "S3", "S4"]:
            sr = r["strategies"].get(sid, {})
            for chunk in sr.get("top_chunks", []):
                ws3.append([
                    r.get("q_id", ""),
                    sid,
                    chunk.get("rank", ""),
                    chunk.get("doc_id", ""),
                    "",
                    (chunk.get("text", "") or "")[:200],
                ])

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


def main():
    with open(QUESTIONS_FILE) as f:
        questions = json.load(f)

    done_ids, all_results = load_checkpoint()
    remaining = [q for q in questions if q["q_id"] not in done_ids]

    print(f"Phase 1 Evaluation — HAE-RAG-CHUNK-v4.0")
    print(f"Config: k=10, fixed-budget, 1800 tokens")
    print(f"Total: {len(questions)} | Done: {len(done_ids)} | Remaining: {len(remaining)}")
    print(f"Estimated time: ~{len(remaining)*90//60} minutes\n")

    for i, q in enumerate(remaining, 1):
        n = len(done_ids)
        print(f"[{n+1:02d}/{len(questions)}] {q['q_id']} [{q['category']}]", end=" ", flush=True)
        t0 = time.time()

        for attempt in range(3):
            try:
                res = run_question(q)
                elapsed = round(time.time() - t0, 1)
                hits = ""
                for sid in STRATEGIES:
                    sr = res["strategies"].get(sid, {})
                    h5 = sr.get("hit_at_5")
                    hits += f"{sid}:{'Y' if h5==1 else ('-' if h5 is None else 'N')} "
                print(f"{elapsed}s | {hits}")
                all_results.append(res)
                done_ids.add(res["q_id"])
                save_checkpoint(done_ids, all_results)
                break
            except Exception as e:
                if attempt < 2:
                    print(f"\n  Retry {attempt+1}/2: {e}", end=" ")
                    time.sleep(10)
                else:
                    print(f"\n  FAILED after 3 attempts: {e}")

        time.sleep(1)

    final_path = f"{RESULTS_DIR}/phase1_final.json"
    with open(final_path, "w") as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)

    excel_path = f"{RESULTS_DIR}/phase1_results.xlsx"
    generate_excel(all_results, excel_path)

    print(f"\n{'='*60}")
    print(f"PHASE 1 COMPLETE: {len(all_results)}/{len(questions)} questions")
    print(f"\nFINAL METRICS:")
    print(f"{'Strategy':<6} {'Hit@5':>8} {'Hit@3':>8} {'Hit@1':>8} {'MRR':>8}")
    print("-" * 38)

    for sid in STRATEGIES:
        h5 = [r["strategies"][sid]["hit_at_5"] for r in all_results
              if r["strategies"].get(sid) and r["strategies"][sid]["hit_at_5"] is not None]
        h3 = [r["strategies"][sid].get("hit_at_3", 0) for r in all_results
              if r["strategies"].get(sid) and r["strategies"][sid].get("hit_at_3") is not None]
        h1 = [r["strategies"][sid].get("hit_at_1", 0) for r in all_results
              if r["strategies"].get(sid) and r["strategies"][sid].get("hit_at_1") is not None]
        mrr = [r["strategies"][sid]["mrr"] for r in all_results
               if r["strategies"].get(sid) and r["strategies"][sid]["mrr"] is not None]
        if h5:
            print(f"{sid:<6} {sum(h5)/len(h5)*100:>7.1f}% {sum(h3)/len(h3)*100 if h3 else 0:>7.1f}% "
                  f"{sum(h1)/len(h1)*100 if h1 else 0:>7.1f}% {sum(mrr)/len(mrr) if mrr else 0:>8.3f}")

    print(f"\nResults: {final_path}")
    print(f"Excel:   {excel_path}")


if __name__ == "__main__":
    main()
