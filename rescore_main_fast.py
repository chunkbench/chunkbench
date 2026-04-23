#!/usr/bin/env python3
"""
Fast rescore for Phase 1 and Phase 2 main result files.

Strategy:
- S1/S2/S3: extract doc_ids from existing top_chunks (no API call)
- S4: re-query to get full retrieved_chunks with doc_ids (~90 calls, ~6 min)
- B0: no retrieval, scores = None

Then rescore all with doc-level hit detection.
Checkpoint-based, resume with --resume.
"""

import json, time, requests, sys, argparse, openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

API_BASE = "http://127.0.0.1:8007"
DATA_DIR = Path("/www/wwwroot/chunkbench.tech/data")
RESULTS  = Path("/www/wwwroot/chunkbench.tech/data/results")

QUESTION_FILE = DATA_DIR / "questions_v1_FROZEN.json"
PHASE1_FILE   = RESULTS / "phase1" / "phase1_final.json"
PHASE2_FILE   = RESULTS / "phase2" / "phase2_final.json"
CHECKPOINT    = RESULTS / "rescore_fast_checkpoint.json"

STRATS = ["S1","S2","S3","S4","B0"]
QUERY_PARAMS = {
    "k": 10, "context_mode": "fixed-budget",
    "context_budget_tokens": 1800, "temperature": 0.0,
    "max_tokens": 512, "dedup": True,
}
PHASE1_CFG = {"S1":"size=512,overlap=50(10%)","S2":"size=512,overlap=50(10%)",
              "S3":"percentile/95","S4":"proposition/default","B0":"no retrieval"}
PHASE2_CFG = {"S1":"size=512,overlap=102(20%)","S2":"size=1024,overlap=102(10%)",
              "S3":"threshold=85,max=2000","S4":"proposition/default","B0":"no retrieval"}


def _norm(d):
    if not d: return d
    if d.endswith(".pdf"): d = d[:-4]
    return d.replace("_shallow","").replace("_deep","")

def doc_hit(src, doc_ids, k):
    if not src or not doc_ids: return 0
    s = _norm(src)
    return int(any(_norm(x)==s for x in doc_ids[:k]))

def doc_mrr(src, doc_ids):
    if not src or not doc_ids: return 0.0
    s = _norm(src)
    for i,x in enumerate(doc_ids):
        if _norm(x)==s: return 1.0/(i+1)
    return 0.0

def call_query(question, mode, strategies):
    r = requests.post(f"{API_BASE}/query",
                      json={"question": question, "mode": mode,
                            "strategies": strategies, **QUERY_PARAMS},
                      timeout=300)
    r.raise_for_status()
    return r.json()["results"]


def run(resume=False):
    questions  = {q["q_id"]: q for q in json.loads(QUESTION_FILE.read_text())}
    phase1_map = {r["q_id"]: r for r in json.loads(PHASE1_FILE.read_text())}
    phase2_map = {r["q_id"]: r for r in json.loads(PHASE2_FILE.read_text())}

    # S4 doc_ids cache: {qid: {phase: [doc_ids]}}
    s4_cache = {}
    done_ids = set()

    if resume and CHECKPOINT.exists():
        cp = json.loads(CHECKPOINT.read_text())
        done_ids = set(cp.get("done_ids",[]))
        s4_cache = cp.get("s4_cache",{})
        print(f"Resumed: {len(done_ids)} done")

    all_ids   = sorted(set(phase1_map)|set(phase2_map))
    remaining = [qid for qid in all_ids if qid not in done_ids]
    total     = len(all_ids)

    print(f"Total: {total} | Remaining S4 queries: {len(remaining)}")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    for n, qid in enumerate(remaining, len(done_ids)+1):
        q = questions.get(qid)
        if not q:
            print(f"[{n:03d}/{total}] {qid}: not in question file, skip")
            done_ids.add(qid); continue

        src = q.get("source_doc","")
        print(f"[{n:03d}/{total}] {qid}  src={src}", flush=True)
        t0 = time.time()

        for attempt in range(3):
            try:
                p1 = call_query(q["question"], "phase1", ["S4"])
                p2 = call_query(q["question"], "phase2", ["S4"])
                s4_cache[qid] = {
                    "phase1": list(dict.fromkeys(c["doc_id"] for c in p1["S4"]["retrieved_chunks"])),
                    "phase2": list(dict.fromkeys(c["doc_id"] for c in p2["S4"]["retrieved_chunks"])),
                    "p1_top3": [{"rank":c["rank"],"doc_id":c["doc_id"],
                                 "distance":round(c.get("distance",0),4),
                                 "text":c["text"][:400]} for c in p1["S4"]["retrieved_chunks"][:3]],
                    "p2_top3": [{"rank":c["rank"],"doc_id":c["doc_id"],
                                 "distance":round(c.get("distance",0),4),
                                 "text":c["text"][:400]} for c in p2["S4"]["retrieved_chunks"][:3]],
                    "p1_answer": p1["S4"].get("answer",""),
                    "p2_answer": p2["S4"].get("answer",""),
                    "p1_ctx":    p1["S4"].get("context_tokens_est",0),
                    "p2_ctx":    p2["S4"].get("context_tokens_est",0),
                }
                elapsed = round(time.time()-t0, 1)
                h1 = doc_hit(src, s4_cache[qid]["phase1"], 5)
                h2 = doc_hit(src, s4_cache[qid]["phase2"], 5)
                print(f"  ok ({elapsed}s) P1_S4={h1} P2_S4={h2}", flush=True)
                break
            except Exception as e:
                if attempt<2: time.sleep(10)
                else: print(f"  SKIP: {e}", flush=True)

        done_ids.add(qid)
        CHECKPOINT.write_text(json.dumps({"done_ids":list(done_ids),"s4_cache":s4_cache},
                                         ensure_ascii=False))
        time.sleep(3)

    # ── Now rescore everything ───────────────────────────────────────────────
    print("\nRescoring all results...")

    for phase_map, phase_key, phase_label in [
        (phase1_map, "phase1", "P1"),
        (phase2_map, "phase2", "P2"),
    ]:
        for qid, entry in phase_map.items():
            q   = questions.get(qid, {})
            src = q.get("source_doc","")
            entry["source_doc"] = src

            strats_entry = entry.get("strategies", {})

            for sid in STRATS:
                sr = strats_entry.get(sid, {})
                if not sr: continue

                if sid == "B0":
                    sr.update({"hit_at_1":None,"hit_at_3":None,"hit_at_5":None,"mrr":None})
                    continue

                if sid == "S4":
                    cache = s4_cache.get(qid,{})
                    doc_ids = cache.get(phase_key, [])
                    top3    = cache.get(f"{phase_key[:2]}_{phase_key[5:]}_top3" if False else
                                        ("p1_top3" if phase_key=="phase1" else "p2_top3"), [])
                    answer  = cache.get("p1_answer" if phase_key=="phase1" else "p2_answer", sr.get("answer",""))
                    ctx     = cache.get("p1_ctx"    if phase_key=="phase1" else "p2_ctx",    sr.get("context_tokens",0))
                    sr["retrieved_doc_ids"] = doc_ids
                    sr["top_chunks"]        = top3
                    sr["answer"]            = answer
                    sr["context_tokens"]    = ctx
                else:
                    # Extract doc_ids from existing top_chunks
                    top3    = sr.get("top_chunks", [])
                    doc_ids = list(dict.fromkeys(c.get("doc_id","") for c in top3 if c.get("doc_id")))
                    sr["retrieved_doc_ids"] = doc_ids

                doc_ids = sr.get("retrieved_doc_ids",[])
                sr["hit_at_1"] = doc_hit(src, doc_ids, 1)
                sr["hit_at_3"] = doc_hit(src, doc_ids, 3)
                sr["hit_at_5"] = doc_hit(src, doc_ids, 5)
                sr["mrr"]      = doc_mrr(src, doc_ids)

    return list(phase1_map.values()), list(phase2_map.values())


def summary(out, label):
    print(f"\n--- {label} ---")
    for sid in STRATS:
        scored = [r for r in out
                  if r.get("strategies",{}).get(sid,{}).get("hit_at_5") is not None
                  and sid != "B0"]
        if not scored: continue
        h5  = [r["strategies"][sid]["hit_at_5"] for r in scored]
        h1  = [r["strategies"][sid]["hit_at_1"] for r in scored]
        mrr = [r["strategies"][sid]["mrr"]       for r in scored]
        print(f"  {sid}: Hit@1={round(sum(h1)/len(h1)*100,1)}%  "
              f"Hit@5={round(sum(h5)/len(h5)*100,1)}%  "
              f"MRR={round(sum(mrr)/len(mrr),3)}  n={len(h5)}")


GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BOLD   = Font(bold=True)


def make_excel(out, path, label, cfg):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Summary"
    ws.append([label])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append(["Hit detection: doc-level (source_doc in top-k retrieved doc_ids)"])
    ws.append([])
    ws.append(["Strategy","Config","N","Hit@1(%)","Hit@3(%)","Hit@5(%)","MRR"])
    for cell in ws[5]: cell.font = BOLD
    for sid in STRATS:
        scored = [r for r in out
                  if r.get("strategies",{}).get(sid,{}).get("hit_at_5") is not None
                  and sid != "B0"]
        if not scored:
            ws.append([sid, cfg.get(sid,""), 0, None, None, None, None]); continue
        h5 = [r["strategies"][sid]["hit_at_5"] for r in scored]
        h3 = [r["strategies"][sid]["hit_at_3"] for r in scored]
        h1 = [r["strategies"][sid]["hit_at_1"] for r in scored]
        m  = [r["strategies"][sid]["mrr"]       for r in scored]
        v5 = round(sum(h5)/len(h5)*100,1)
        ws.append([sid, cfg.get(sid,""), len(h5),
                   round(sum(h1)/len(h1)*100,1), round(sum(h3)/len(h3)*100,1),
                   v5, round(sum(m)/len(m),3)])
        ws.cell(ws.max_row,6).fill = GREEN if v5>=40 else RED if v5<15 else YELLOW

    ws2 = wb.create_sheet("Per-Question")
    hdr = ["q_id","source_doc","category","needle_type","question","gold_span","gold_answer"]
    for sid in STRATS: hdr += [f"{sid}_hit5",f"{sid}_hit3",f"{sid}_mrr",f"{sid}_docs"]
    for sid in STRATS: hdr.append(f"{sid}_answer")
    ws2.append(hdr)
    for cell in ws2[1]: cell.font = BOLD
    for r in out:
        row = [r["q_id"], r.get("source_doc",""), r.get("category",""), r.get("needle_type",""),
               r.get("question","")[:150], r.get("gold_span","")[:150], r.get("gold_answer","")[:100]]
        for sid in STRATS:
            sr   = r.get("strategies",{}).get(sid,{})
            docs = ", ".join(sr.get("retrieved_doc_ids",[]))
            row += [sr.get("hit_at_5"), sr.get("hit_at_3"), sr.get("mrr"), docs]
        for sid in STRATS:
            row.append(r.get("strategies",{}).get(sid,{}).get("answer","")[:250])
        ws2.append(row)
        rn = ws2.max_row
        for i, h in enumerate(hdr):
            if h.endswith("_hit5"):
                c = ws2.cell(rn, i+1)
                if c.value==1: c.fill=GREEN
                elif c.value==0: c.fill=RED

    for s in [ws, ws2]:
        for col in s.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            s.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2,45)
    wb.save(str(path))
    print(f"  Excel: {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true")
    args = parser.parse_args()

    try:
        r = requests.get(f"{API_BASE}/health", timeout=10)
        print(f"API: {r.json()}")
    except Exception as e:
        print(f"API not reachable: {e}"); sys.exit(1)

    p1_out, p2_out = run(resume=args.resume)

    PHASE1_FILE.write_text(json.dumps(p1_out, indent=2, ensure_ascii=False))
    PHASE2_FILE.write_text(json.dumps(p2_out, indent=2, ensure_ascii=False))
    print("\nJSON saved.")

    summary(p1_out, "Phase 1")
    summary(p2_out, "Phase 2")

    make_excel(p1_out, RESULTS/"phase1"/"phase1_results.xlsx",
               "Phase 1 — 90 questions (doc-level hit)", PHASE1_CFG)
    make_excel(p2_out, RESULTS/"phase2"/"phase2_results.xlsx",
               "Phase 2 — 90 questions (doc-level hit)", PHASE2_CFG)

    print(f"\nDone: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
