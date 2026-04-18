#!/usr/bin/env python3
import json, os, sys, time, re, shutil, subprocess, requests
from pathlib import Path
from datetime import datetime

API_BASE    = "http://127.0.0.1:8007"
INDEX_BASE  = Path("indexes")
RESULTS_DIR = Path("data/results/phase2")
QUESTIONS   = "data/questions_v1_FROZEN.json"
BEST_CFG    = "data/results/phase2_tuning/best_configs.json"
CHECKPOINT  = RESULTS_DIR / "checkpoint.json"
STRATEGIES  = ["S1", "S2", "S3", "S4", "B0"]

EVAL_PARAMS = {
    "k": 10, "context_mode": "fixed-budget",
    "context_budget_tokens": 1800, "temperature": 0.0, "dedup": True,
}

RESULTS_DIR.mkdir(parents=True, exist_ok=True)


def swap_best_configs():
    with open(BEST_CFG) as f:
        best = json.load(f)

    for sid in ["S1", "S2", "S3"]:
        label = best[sid]["label"]
        src = INDEX_BASE / f"chroma_{sid.lower()}_{label}"
        dst = INDEX_BASE / f"chroma_{sid.lower()}"
        if src.exists():
            if dst.exists():
                shutil.rmtree(str(dst))
            shutil.copytree(str(src), str(dst))
            print(f"  {sid}: swapped to {label}", flush=True)
        else:
            print(f"  {sid}: WARNING — {src} not found, keeping current", flush=True)

    subprocess.run(["systemctl", "restart", "chunkbench"], check=False)
    time.sleep(15)
    r = requests.get(f"{API_BASE}/health", timeout=10)
    print(f"  API: {r.json()}", flush=True)


def run_eval():
    with open(QUESTIONS) as f:
        questions = json.load(f)

    done_ids, results = set(), []
    if CHECKPOINT.exists():
        with open(CHECKPOINT) as f:
            cp = json.load(f)
        done_ids = set(cp.get("done_ids", []))
        results = cp.get("results", [])

    remaining = [q for q in questions if q["q_id"] not in done_ids]
    print(f"\nPhase 2 Eval: {len(questions)} total | {len(done_ids)} done | {len(remaining)} remaining", flush=True)

    for q in remaining:
        n = len(done_ids) + 1
        print(f"  [{n:02d}/{len(questions)}] {q['q_id']}", end=" ", flush=True)
        t0 = time.time()

        payload = {
            "questions": [{
                "q_id": q["q_id"], "question": q["question"],
                "category": q.get("category", ""), "needle_type": q.get("needle_type", "corpus"),
                "gold_answer": q.get("gold_answer", ""), "gold_span": q.get("gold_span", ""),
                "gold_span_2": q.get("gold_span_2"), "source_doc": q.get("source_doc", ""),
                "difficulty": q.get("difficulty", ""),
            }],
            "strategies": STRATEGIES,
            **EVAL_PARAMS,
        }

        for attempt in range(3):
            try:
                r = requests.post(f"{API_BASE}/evaluate", json=payload, timeout=300)
                res = r.json()["results"][0]
                elapsed = round(time.time() - t0, 1)
                hits = " ".join(f"{s}:{'Y' if res['strategies'].get(s,{}).get('hit_at_5')==1 else 'N'}"
                               for s in ["S1","S2","S3","S4"])
                print(f"{elapsed}s | {hits}", flush=True)
                results.append(res)
                done_ids.add(res["q_id"])
                with open(CHECKPOINT, "w") as f:
                    json.dump({"done_ids": list(done_ids), "results": results}, f)
                break
            except Exception as e:
                if attempt < 2:
                    time.sleep(10)
                else:
                    print(f"FAILED: {e}", flush=True)
        time.sleep(1)

    with open(RESULTS_DIR / "phase2_final.json", "w") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    return results


def generate_excel(results):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
    def clean(s):
        return ILLEGAL.sub('', s) if isinstance(s, str) else s

    with open(BEST_CFG) as f:
        best = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["HAE-RAG Phase 2 Results — v4.0 (Optimized)"])
    ws.append(["Config: k=10, fixed-budget 1800tok, best per-strategy params"])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    ws.append(["Strategy", "Best Config", "Hit@5 (%)", "Hit@3 (%)", "Hit@1 (%)", "MRR", "Avg Ctx Tokens"])
    for cell in ws[5]: cell.font = Font(bold=True)

    for sid in STRATEGIES:
        h5  = [r["strategies"][sid]["hit_at_5"] for r in results if r["strategies"].get(sid) and r["strategies"][sid]["hit_at_5"] is not None]
        h3  = [r["strategies"][sid].get("hit_at_3",0) for r in results if r["strategies"].get(sid) and r["strategies"][sid].get("hit_at_3") is not None]
        h1  = [r["strategies"][sid].get("hit_at_1",0) for r in results if r["strategies"].get(sid) and r["strategies"][sid].get("hit_at_1") is not None]
        mrr = [r["strategies"][sid]["mrr"] for r in results if r["strategies"].get(sid) and r["strategies"][sid]["mrr"] is not None]
        ctx = [r["strategies"][sid].get("context_tokens",0) for r in results if r["strategies"].get(sid)]
        cfg_label = best.get(sid, {}).get("label", "default")
        if h5:
            ws.append([sid, cfg_label,
                       round(sum(h5)/len(h5)*100, 1),
                       round(sum(h3)/len(h3)*100, 1) if h3 else "N/A",
                       round(sum(h1)/len(h1)*100, 1) if h1 else "N/A",
                       round(sum(mrr)/len(mrr), 3) if mrr else "N/A",
                       round(sum(ctx)/len(ctx)) if ctx else 0])
        else:
            ws.append([sid, cfg_label, "N/A", "N/A", "N/A", "N/A", 0])

    ws.append([])
    ws.append(["Category Breakdown (Hit@5 %)"])
    ws.append(["Category", "S1", "S2", "S3", "S4"])
    for cat in ["single_fact", "numerical", "multi_hop", "negation", "sentinel"]:
        row = [cat]
        for sid in ["S1", "S2", "S3", "S4"]:
            qs = [r for r in results if r["category"] == cat]
            h5 = [r["strategies"][sid]["hit_at_5"] for r in qs if r["strategies"].get(sid) and r["strategies"][sid]["hit_at_5"] is not None]
            row.append(f"{sum(h5)/len(h5)*100:.0f}%" if h5 else "N/A")
        ws.append(row)

    ws2 = wb.create_sheet("Per-Question")
    headers = ["q_id", "question", "category", "gold_answer",
               "S1_hit5", "S1_answer", "S2_hit5", "S2_answer",
               "S3_hit5", "S3_answer", "S4_hit5", "S4_answer", "B0_answer"]
    ws2.append(headers)
    for cell in ws2[1]: cell.font = Font(bold=True)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for r in results:
        row = [r.get("q_id",""), clean(r.get("question",""))[:100],
               r.get("category",""), clean(r.get("gold_answer",""))[:80]]
        for sid in ["S1", "S2", "S3", "S4"]:
            sr = r["strategies"].get(sid, {})
            row.extend([sr.get("hit_at_5", ""), clean((sr.get("answer", "") or ""))[:150]])
        row.append(clean((r["strategies"].get("B0", {}).get("answer", "") or ""))[:150])
        ws2.append(row)
        rn = ws2.max_row
        for col in [5, 7, 9, 11]:
            cell = ws2.cell(row=rn, column=col)
            if cell.value == 1: cell.fill = green
            elif cell.value == 0: cell.fill = red

    out = RESULTS_DIR / "phase2_results.xlsx"
    wb.save(str(out))
    print(f"Excel: {out}", flush=True)


def main():
    print("="*60, flush=True)
    print("PHASE 2 FULL EVALUATION", flush=True)
    print("="*60, flush=True)

    print("\nSwapping to best configs...", flush=True)
    swap_best_configs()

    print("\nRunning 90-question evaluation...", flush=True)
    results = run_eval()

    print("\nGenerating Excel...", flush=True)
    generate_excel(results)

    print(f"\n{'='*60}", flush=True)
    print("PHASE 2 COMPLETE", flush=True)
    for sid in ["S1", "S2", "S3", "S4"]:
        h5 = [r["strategies"][sid]["hit_at_5"] for r in results if r["strategies"].get(sid) and r["strategies"][sid]["hit_at_5"] is not None]
        if h5:
            print(f"  {sid}: Hit@5={sum(h5)/len(h5)*100:.1f}%", flush=True)

    print("\nRestoring Phase 1 indexes...", flush=True)
    for sid in ["S1", "S2", "S3"]:
        backup = INDEX_BASE / f"chroma_{sid.lower()}_phase1_backup"
        prod = INDEX_BASE / f"chroma_{sid.lower()}"
        if backup.exists():
            if prod.exists(): shutil.rmtree(str(prod))
            shutil.copytree(str(backup), str(prod))
    subprocess.run(["systemctl", "restart", "chunkbench"], check=False)
    print("Done — Phase 1 indexes restored.", flush=True)


if __name__ == "__main__":
    main()
