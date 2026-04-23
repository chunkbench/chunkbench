#!/usr/bin/env python3
"""
Rescores all existing result JSONs using doc-level hit detection.

Old (wrong): gold_span substring present in chunk text  →  fails for S4 (paraphrases)
New (correct): source_doc present in top-k retrieved_doc_ids

Works entirely from existing JSON data — no API calls needed.
Regenerates all Excel reports.
"""

import json, re, openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

RESULTS = Path("/www/wwwroot/chunkbench.tech/data/results")

STRATS = ["S1", "S2", "S3", "S4", "B0"]

PHASE2_CFG = {
    "S1": "size=512, overlap=102 (20%)",
    "S2": "size=1024, overlap=102 (10%)",
    "S3": "threshold=85, max=2000",
    "S4": "proposition / default",
    "B0": "no retrieval",
}
PHASE1_CFG = {
    "S1": "size=512, overlap=50 (10%)",
    "S2": "size=512, overlap=50 (10%)",
    "S3": "percentile/95",
    "S4": "proposition / default",
    "B0": "no retrieval",
}


def clean(s):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s) if isinstance(s, str) else s


def normalize_doc_id(doc_id):
    """Strip .pdf suffix and _shallow/_deep suffixes for comparison."""
    if not doc_id:
        return doc_id
    d = doc_id
    if d.endswith(".pdf"):
        d = d[:-4]
    d = d.replace("_shallow", "").replace("_deep", "")
    return d


def doc_hit(source_doc, doc_ids, k):
    if not source_doc or not doc_ids:
        return 0
    src = normalize_doc_id(source_doc)
    return int(any(normalize_doc_id(d) == src for d in doc_ids[:k]))


def doc_mrr(source_doc, doc_ids):
    if not source_doc or not doc_ids:
        return 0.0
    src = normalize_doc_id(source_doc)
    for i, d in enumerate(doc_ids):
        if normalize_doc_id(d) == src:
            return 1.0 / (i + 1)
    return 0.0


def rescore_entry(result, phase):
    source_doc = result.get("source_doc", "")
    phase_data = result.get(phase, {})
    for sid in STRATS:
        sr = phase_data.get(sid)
        if not sr:
            continue
        if sid == "B0":
            sr["hit_at_1"] = None
            sr["hit_at_3"] = None
            sr["hit_at_5"] = None
            sr["mrr"]      = None
            continue
        doc_ids = sr.get("retrieved_doc_ids", [])
        if not doc_ids:
            continue
        sr["hit_at_1"] = doc_hit(source_doc, doc_ids, 1)
        sr["hit_at_3"] = doc_hit(source_doc, doc_ids, 3)
        sr["hit_at_5"] = doc_hit(source_doc, doc_ids, 5)
        sr["mrr"]      = doc_mrr(source_doc, doc_ids)


def print_summary(out, label, phase):
    print(f"\n--- {label} ({phase}) ---")
    for sid in STRATS:
        scored = [r for r in out
                  if r.get(phase, {}).get(sid) and
                  r[phase][sid].get("hit_at_5") is not None and
                  sid != "B0"]
        if not scored:
            continue
        h5  = [r[phase][sid]["hit_at_5"] for r in scored]
        h1  = [r[phase][sid]["hit_at_1"] for r in scored]
        mrr = [r[phase][sid]["mrr"]       for r in scored]
        print(f"  {sid}: Hit@1={round(sum(h1)/len(h1)*100,1)}%  "
              f"Hit@5={round(sum(h5)/len(h5)*100,1)}%  "
              f"MRR={round(sum(mrr)/len(mrr),3)}  n={len(h5)}")


# ── Excel helpers ────────────────────────────────────────────────────────────

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BOLD   = Font(bold=True)


def auto_width(ws):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 45)


def summary_sheet(ws, out, phase, cfg, title):
    ws.append([title])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append(["Hit detection: doc-level (source_doc ∈ top-k retrieved doc_ids)"])
    ws.append([])
    ws.append(["Strategy", "Config", "N", "Hit@1(%)", "Hit@3(%)", "Hit@5(%)", "MRR"])
    for cell in ws[5]:
        cell.font = BOLD
    for sid in STRATS:
        scored = [r for r in out
                  if r.get(phase, {}).get(sid) and
                  r[phase][sid].get("hit_at_5") is not None and
                  sid != "B0"]
        if not scored:
            ws.append([sid, cfg.get(sid, ""), 0, None, None, None, None])
            continue
        h5  = [r[phase][sid]["hit_at_5"] for r in scored]
        h3  = [r[phase][sid]["hit_at_3"] for r in scored]
        h1  = [r[phase][sid]["hit_at_1"] for r in scored]
        m   = [r[phase][sid]["mrr"]       for r in scored]
        v5  = round(sum(h5) / len(h5) * 100, 1)
        ws.append([sid, cfg.get(sid, ""), len(h5),
                   round(sum(h1) / len(h1) * 100, 1),
                   round(sum(h3) / len(h3) * 100, 1),
                   v5,
                   round(sum(m) / len(m), 3)])
        ws.cell(ws.max_row, 6).fill = (GREEN if v5 >= 40 else RED if v5 < 15 else YELLOW)


def perq_sheet(ws, out, phase):
    hdr = ["q_id", "source_doc", "question", "gold_span", "gold_answer"]
    for sid in STRATS:
        hdr += [f"{sid}_hit5", f"{sid}_hit3", f"{sid}_mrr", f"{sid}_ctx_tok", f"{sid}_docs"]
    for sid in STRATS:
        hdr.append(f"{sid}_answer")
    ws.append(hdr)
    for cell in ws[1]:
        cell.font = BOLD
    for r in out:
        row = [r["q_id"], r.get("source_doc", ""),
               clean(r.get("question", ""))[:150],
               clean(r.get("gold_span", ""))[:150],
               clean(r.get("gold_answer", ""))[:100]]
        for sid in STRATS:
            sr   = r.get(phase, {}).get(sid, {})
            docs = ", ".join(sr.get("retrieved_doc_ids", []))
            row += [sr.get("hit_at_5"), sr.get("hit_at_3"),
                    sr.get("mrr"), sr.get("context_tokens"), docs]
        for sid in STRATS:
            row.append(clean(r.get(phase, {}).get(sid, {}).get("answer", ""))[:250])
        ws.append(row)
        rn = ws.max_row
        for i, h in enumerate(hdr):
            if h.endswith("_hit5"):
                c = ws.cell(rn, i + 1)
                if c.value == 1:
                    c.fill = GREEN
                elif c.value == 0:
                    c.fill = RED


# ── Process each dataset ─────────────────────────────────────────────────────

def process_new_questions():
    rfile = RESULTS / "new_questions_eval" / "new_questions_results_v2.json"
    out   = json.loads(rfile.read_text(encoding="utf-8"))
    for r in out:
        rescore_entry(r, "phase1")
        rescore_entry(r, "phase2")
    rfile.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print_summary(out, "New Questions (65q)", "phase1")
    print_summary(out, "New Questions (65q)", "phase2")

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Phase1 Summary"
    summary_sheet(ws, out, "phase1", PHASE1_CFG, "New Questions — Phase 1")
    ws2 = wb.create_sheet("Phase2 Summary")
    summary_sheet(ws2, out, "phase2", PHASE2_CFG, "New Questions — Phase 2")
    ws3 = wb.create_sheet("Phase1 Per-Question")
    perq_sheet(ws3, out, "phase1")
    ws4 = wb.create_sheet("Phase2 Per-Question")
    perq_sheet(ws4, out, "phase2")
    for s in [ws, ws2, ws3, ws4]:
        auto_width(s)
    path = RESULTS / "new_questions_eval" / "new_questions_eval_v2.xlsx"
    wb.save(str(path))
    print(f"  Excel: {path}")


def process_sentinel():
    rfile = RESULTS / "sentinel_eval" / "sentinel_results.json"
    out   = json.loads(rfile.read_text(encoding="utf-8"))
    for r in out:
        rescore_entry(r, "phase1")
        rescore_entry(r, "phase2")
    rfile.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print_summary(out, "Sentinel (20q)", "phase1")
    print_summary(out, "Sentinel (20q)", "phase2")

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Phase1 Summary"
    summary_sheet(ws, out, "phase1", PHASE1_CFG, "Sentinel — Phase 1")
    ws2 = wb.create_sheet("Phase2 Summary")
    summary_sheet(ws2, out, "phase2", PHASE2_CFG, "Sentinel — Phase 2")
    ws3 = wb.create_sheet("Phase1 Per-Question")
    perq_sheet(ws3, out, "phase1")
    ws4 = wb.create_sheet("Phase2 Per-Question")
    perq_sheet(ws4, out, "phase2")
    for s in [ws, ws2, ws3, ws4]:
        auto_width(s)
    path = RESULTS / "sentinel_eval" / "sentinel_eval.xlsx"
    wb.save(str(path))
    print(f"  Excel: {path}")


def process_ek_cat1():
    rfile = RESULTS / "ek_cat1_eval" / "ek_cat1_results.json"
    out   = json.loads(rfile.read_text(encoding="utf-8"))
    for r in out:
        rescore_entry(r, "phase2")
    rfile.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print_summary(out, "EK Cat1 (13q)", "phase2")

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Phase2 Summary"
    summary_sheet(ws, out, "phase2", PHASE2_CFG, "EK Cat1 — Phase 2")
    ws2 = wb.create_sheet("Phase2 Per-Question")
    perq_sheet(ws2, out, "phase2")
    for s in [ws, ws2]:
        auto_width(s)
    path = RESULTS / "ek_cat1_eval" / "ek_cat1_eval.xlsx"
    wb.save(str(path))
    print(f"  Excel: {path}")


def main():
    print(f"Rescoring with doc-level hit detection...")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    process_new_questions()
    process_sentinel()
    process_ek_cat1()
    print(f"\nDone: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
